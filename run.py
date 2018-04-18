import json
from multiprocessing import cpu_count

from bottle_errorsrest import ErrorsRestPlugin
from bottle import (
    JSONPlugin,
    get,
    static_file,
    install,
    request,
    HTTPError,
    run)
from openpyxl import load_workbook
from pymongo import MongoClient

db = MongoClient(connect=False)['flightlens']

labels = [
    "_id", "DATE", "DEP", "DEP_TIME", "DEP_LOCAL_TIME", "ARR", "ARR_TIME",
    "ARR_LOCAL_TIME", "BaseIataCode", "LOF_ID"]


def load_excel(filename):
    # TODO: create indexes

    wb = load_workbook(filename)
    worksheet = wb[wb.sheetnames[0]]
    for row in worksheet:
        if not isinstance(row[0].value, int):
            # skip headers
            continue

        db.flight.insert({
            labels[n]: row[n].value for n in range(len(labels))
            })


@get('/static/<path:path>')
def static(path):
    return static_file(path, 'static')


@get('/')
def index():
    return static('index.html')


@get('/flights')
def flights():
    # TODO: use https://github.com/kianxineki/bottle-marshmallow
    def secure_int(value):
        if isinstance(value, int):
            return value

        if hasattr(value, 'isdigit'):
            return int(value)

        raise HTTPError('invalid number')

    sort = request.query.get("sort", "_id")
    limit = secure_int(request.query.get("limit", 10))
    offset = secure_int(request.query.get("offset", 0))
    order = request.query.get("order", 'asc')
    search = request.query.get("search")

    if sort not in labels:
        print(labels, sort)
        raise HTTPError(400, 'invalid label')

    if order not in ['asc', 'desc']:
        raise HTTPError(400, 'invalid sort')

    if offset < 0:
        raise HTTPError(400, 'invalid offset')

    if search:
        search = search.upper()
        search = {
            '$or': [
                {'BaseIataCode': {'$regex': search}},
                {'DEP': {'$regex': search}},
                {'ARR': {'$regex': search}},
                ]
        }
    else:
        search = {}
    qry = db.flight.find(search).sort(
        sort, 1 if order == 'asc' else -1).skip(offset).limit(limit)
    return {
        'total': qry.count(),
        'rows': list(qry)}


@get('/flights/<flight_id:int>')
def flight(flight_id):
    result = db.flight.find_one({'_id': flight_id})
    if not result:
        raise HTTPError(404, "the flight does not exist")
    return result


class JSONEncoder(json.JSONEncoder):
    def default(self, o):
        try:
            json.JSONEncoder.default(self, o)
        except TypeError:
            return str(o)


if __name__ == "__main__":
    if not MongoClient()['flightlens'].flight.count():
        load_excel('FlightLegs-2017-07-24.xlsx')
    # install plugins
    install(JSONPlugin(json_dumps=JSONEncoder().encode))
    install(ErrorsRestPlugin())
    run(port=4422,
        host="0.0.0.0",
        server="gunicorn",
        workers=2 * cpu_count() + 1,
        worker_class="egg:meinheld#gunicorn_worker",
        quiet=True,
        debug=False,
        timeout=600)
